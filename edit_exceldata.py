
#pip install xlrdをする
#pip install openpyxlをする


import pandas as pd
from datetime import datetime
import openpyxl
import subprocess

#センサ番号から部屋名を抽出する
def ExtractRoomName(SensorNumber):
	df  = pd.read_excel('datasheet.xlsx',sheet_name = 1)
	#print(df)
	df1 = df.query('センサ番号 == @SensorNumber')
	#print(df1)
	RoomName = df1.iat[0,2]
	print(RoomName)
	return RoomName

#建物番号から建物名を抽出する
def ExtractBldName(BldNumber):
	df  = pd.read_excel('datasheet.xlsx',sheet_name = 0)
	#print(df)
	df1 = df.query('建物番号 == @BldNumber')
	#print(df1)
	BldName = df1.iat[0,1]
	print(BldName)
	return BldName

#センサ番号から建物番号を抽出する
def ExtractBldNumber(SensorNumber):
	df  = pd.read_excel('datasheet.xlsx',sheet_name = 1)
	#print(df)
	df1 = df.query('センサ番号 == @SensorNumber')
	#print(df1)
	BldNumber = df1.iat[0,0]
	#print(BldNumber)
	return BldNumber


#Excelファイルの最終行に時間と場所を記述する(受信時)
def writeReceiveData(RoomName,sensornumber):
	today = datetime.today().strftime('%Y/%m/%d %H:%M:%S')
	wb = openpyxl.load_workbook('datasheet.xlsx')
	ws = wb.worksheets[2]
	maxRow = ws.max_row + 1
	#print(maxRow)
	ws.cell(row=maxRow,column=1).value= today
	ws.cell(row=maxRow,column=2).value= RoomName
	ws.cell(row=maxRow,column=3).value= sensornumber
	wb.save('datasheet.xlsx')


#Excelファイルの最終行に時間と場所と状態を記述する(送信時)
def writeSendData(BldData,state):
	today = datetime.today().strftime('%Y/%m/%d %H:%M:%S')
	wb = openpyxl.load_workbook('datasheet.xlsx')
	ws = wb.worksheets[3]
	maxRow = ws.max_row + 1
	#print(maxRow)
	ws.cell(row=maxRow,column=1).value= today
	ws.cell(row=maxRow,column=2).value= BldData
	ws.cell(row=maxRow,column=3).value= state
	wb.save('datasheet.xlsx')

#Excelファイルを起動する
def openExcel():
	subprocess.Popen(['start',"datasheet.xlsx"], shell=True)
	


